# api/core/presentacion_excel_export.py
from __future__ import annotations

import re
from io import BytesIO
from typing import Literal

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font
from core.presentacion_export import OrigenAnalisis, etiqueta_origen_analisis

ExcelChartKind = Literal["column", "line", "pie", "none"]


def _clave_orden_periodo(periodo: object) -> tuple[int, float, str]:
    """Clave para ordenar de más antiguo a más reciente (eje X Excel: izquierda → derecha)."""
    if isinstance(periodo, bool):
        return (3, 0.0, str(periodo))
    if isinstance(periodo, (int, float)):
        return (0, float(periodo), "")
    s = str(periodo).strip()
    if not s:
        return (4, 0.0, "")
    head = s.replace("T", " ").strip()
    md = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", head)
    if md:
        y, mo, d = int(md.group(1)), int(md.group(2)), int(md.group(3))
        return (0, float(y * 10000 + mo * 100 + d), "")
    my = re.match(r"^(\d{4})-(\d{1,2})(?:$|[^\d])", head)
    if my:
        y, mo = int(my.group(1)), int(my.group(2))
        return (0, float(y * 10000 + mo * 100 + 1), "")
    m4 = re.match(r"^(\d{4})(?:$|[^\d])", head)
    if m4:
        y = int(m4.group(1))
        return (0, float(y * 10000 + 101), "")
    return (2, 0.0, s.lower())


def _ordenar_puntos_cronologico(
    puntos: list[tuple[object, float, str | None]],
) -> list[tuple[object, float, str | None]]:
    if len(puntos) < 2:
        return list(puntos)
    return sorted(
        puntos,
        key=lambda row: (
            _clave_orden_periodo(row[0]),
            ((row[2] or "").strip()).lower(),
        ),
    )


def _slug_sheet_name(title: str, index: int) -> str:
    raw = re.sub(r"[\[\]\*:\\/\?\]]", " ", title, flags=re.UNICODE).strip()
    raw = re.sub(r"\s+", " ", raw)[:24] or "Grafica"
    return f"{index:02d}_{raw}"[:31]


def _cell_text(s: str | None) -> str:
    if not s:
        return ""
    return str(s).strip()


def construir_xlsx_lote_bytes(
    titulo_portada: str,
    secciones: list[
        tuple[
            str,  # titulo
            str | None,  # subtitulo
            str,  # cuerpo_analisis
            OrigenAnalisis,
            str | None,  # leyenda
            ExcelChartKind,
            list[tuple[str | int | float, float, str | None]],  # periodo, valor, entidad_clave
        ]
    ],
) -> bytes:
    """Un libro, una hoja por sección; gráfico nativo salvo kind ``none``."""
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for idx, (titulo, subtitulo, cuerpo, origen, leyenda, chart_kind, puntos) in enumerate(
        secciones, start=1
    ):
        name = _slug_sheet_name(titulo, idx)
        ws = wb.create_sheet(title=name)

        title_font = Font(name="Calibri", size=14, bold=True)
        body_font = Font(name="Calibri", size=11)

        ws["A1"] = _cell_text(titulo)
        ws["A1"].font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        r = 2
        if subtitulo and str(subtitulo).strip():
            ws[f"A{r}"] = _cell_text(subtitulo)
            ws[f"A{r}"].font = Font(name="Calibri", size=11, italic=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1

        hdr = r + 1
        ws[f"A{hdr}"] = "Período"
        ws[f"B{hdr}"] = "Valor"
        ws[f"C{hdr}"] = "Entidad"
        ws[f"D{hdr}"] = "Categoría"
        for c in range(1, 5):
            ws.cell(row=hdr, column=c).font = Font(bold=True)

        puntos_ord = _ordenar_puntos_cronologico(list(puntos))
        row = hdr + 1
        for periodo, valor, ent in puntos_ord:
            ent_s = (ent or "").strip() if ent else ""
            cat = f"{periodo} — {ent_s}" if ent_s else str(periodo)
            ws.cell(row=row, column=1, value=periodo)
            ws.cell(row=row, column=2, value=float(valor))
            ws.cell(row=row, column=3, value=ent_s or None)
            ws.cell(row=row, column=4, value=cat)
            row += 1

        data_end = row - 1
        if data_end >= hdr + 1 and chart_kind != "none":
            n_start = hdr
            n_end = data_end
            vals = Reference(ws, min_col=2, min_row=n_start, max_row=n_end)
            cats = Reference(ws, min_col=4, min_row=n_start + 1, max_row=n_end)

            if chart_kind == "pie":
                chart = PieChart()
                chart.add_data(vals, titles_from_data=True)
                chart.set_categories(cats)
                chart.title = _cell_text(titulo)[:80] or "Gráfica"
                ws.add_chart(chart, "F2")
            elif chart_kind == "line":
                chart = LineChart()
                chart.add_data(vals, titles_from_data=True)
                chart.set_categories(cats)
                chart.title = _cell_text(titulo)[:80] or "Gráfica"
                chart.y_axis.title = "Valor"
                ws.add_chart(chart, "F2")
            else:
                chart = BarChart()
                chart.type = "col"
                chart.add_data(vals, titles_from_data=True)
                chart.set_categories(cats)
                chart.title = _cell_text(titulo)[:80] or "Gráfica"
                chart.y_axis.title = "Valor"
                ws.add_chart(chart, "F2")

        analisis_row = max(row + 2, hdr + 18)
        etiqueta = etiqueta_origen_analisis(origen)
        ws.cell(row=analisis_row, column=1, value=etiqueta)
        ws.cell(row=analisis_row, column=1).font = Font(bold=True)
        ws.merge_cells(
            start_row=analisis_row + 1,
            start_column=1,
            end_row=analisis_row + 8,
            end_column=8,
        )
        cell = ws.cell(row=analisis_row + 1, column=1, value=_cell_text(cuerpo) or "—")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = body_font

        if leyenda and str(leyenda).strip():
            foot = analisis_row + 10
            ws.cell(row=foot, column=1, value=str(leyenda).strip())
            ws.cell(row=foot, column=1).font = Font(name="Calibri", size=9, italic=True)
            ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=8)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 36

    meta = wb.create_sheet(title="00_Portada", index=0)
    meta["A1"] = _cell_text(titulo_portada) or "AEDMI"
    meta["A1"].font = Font(name="Calibri", size=16, bold=True)
    meta["A3"] = "Exportación con datos y gráficos nativos de Excel (sin imagen PNG)."

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
